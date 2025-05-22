# ---------------------------------------------------------------------------------------------------
# APLICACIÓN COMPLETA - MODELO PREDICTIVO - MAS VERSION OPTIMIZADA FINAL + GESTION NUEVA HEURISTICO
# ---------------------------------------------------------------------------------------------------

import os
import time
import hashlib
import pygetwindow as gw
import numpy as np
import pandas as pd
import streamlit.components.v1 as components
import streamlit as st
import subprocess
import pyautogui
import openpyxl
import os
os.environ['TF_CPP_MIN_LOG_LEVEL'] = '2'
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime
from itertools import cycle
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager
from sklearn.ensemble import RandomForestClassifier
from sklearn.model_selection import train_test_split
from sklearn.metrics import accuracy_score, classification_report
from screeninfo import get_monitors
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import OneHotEncoder
from sklearn.pipeline import Pipeline
from sklearn.compose import ColumnTransformer

tiempo_inicio = time.time()

# --- Configuración general ---
st.set_page_config(page_title="Ruleta - Francisco M", layout="wide")
st.title("🎰 Estrategia Segura | Tu Asistente Predictivo")
tab1 = st.tabs(["🧠 Modelo Predictivo"])[0]

# --- Estado sesión ---
if 'tabla_jugadas' not in st.session_state:
    st.session_state.tabla_jugadas = []

if 'historial' not in st.session_state:
    st.session_state.historial = []
    st.session_state.fechas = []
    st.session_state.data_arr = np.array([], dtype=int)
    st.session_state.total_inicial = 0
    st.session_state.spins_nuevos = 0
    st.session_state.ultima_firma = None
    st.session_state.model = None
    st.session_state.last_col = None
    st.session_state.aciertos = 0
    st.session_state.perdidas = 0
    st.session_state.initialized = False
    st.session_state.matriz_transicion = {}

if 'pausar' not in st.session_state:
    st.session_state.pausar = False

if 'resultado_ruleta' not in st.session_state:
    st.session_state.resultado_ruleta = ""

if 'numeros_no_guardados' not in st.session_state:
    st.session_state.numeros_no_guardados = []

if 'columna_predicha_anterior' not in st.session_state:
    st.session_state.columna_predicha_anterior = None

if 'esperando_numero' not in st.session_state:
    st.session_state.esperando_numero = False

if 'contador_ciclos' not in st.session_state:
    st.session_state.contador_ciclos = 0

if 'modelo_rf_largo' not in st.session_state:
    st.session_state.modelo_rf_largo = None

if 'historial_accuracy' not in st.session_state:
    st.session_state.historial_accuracy = []

# Añade esto junto a tus otros flags de sesión:
if 'simular_restantes' not in st.session_state:
    st.session_state.simular_restantes = 0

if 'reset_progression' not in st.session_state:
    st.session_state.reset_progression = False

if 'correo_enviado_ultima_vez' not in st.session_state:
    st.session_state.correo_enviado_ultima_vez = 0

# --- Constantes ---
NUMEROS_ROJOS = {1,3,5,7,9,12,14,16,18,19,21,23,25,27,30,32,34,36}
NUMEROS_NEGROS = {2,4,6,8,10,11,13,15,17,20,22,24,26,28,29,31,33,35}
VOISINS_DU_ZERO = {22,18,29,7,28,12,35,3,26,0,32,15,19,4,21,2,25}
NUMEROS_ALTOS = set(range(19, 37))
NUMEROS_BAJOS = set(range(1, 19))
DECENAS = {
    1: set(range(1, 13)),
    2: set(range(13, 25)),
    3: set(range(25, 37))
}

def columna_con_mas_voisins(hist):
    col_count = [0, 0, 0]
    for n in hist:
        if int(n) in VOISINS_DU_ZERO:
            col = column_of_number(int(n))
            if col: col_count[col - 1] += 1
    return col_count.index(max(col_count)) + 1

def columna_con_mas_altos(hist):
    col_count = [0, 0, 0]
    for n in hist:
        if int(n) in NUMEROS_ALTOS:
            col = column_of_number(int(n))
            if col: col_count[col - 1] += 1
    return col_count.index(max(col_count)) + 1

def columna_con_mas_bajos(hist):
    col_count = [0, 0, 0]
    for n in hist:
        if int(n) in NUMEROS_BAJOS:
            col = column_of_number(int(n))
            if col: col_count[col - 1] += 1
    return col_count.index(max(col_count)) + 1

def columna_con_mas_decenas(hist):
    col_count = [0, 0, 0]
    for n in hist:
        for dec, rng in DECENAS.items():
            if int(n) in rng:
                col = column_of_number(int(n))
                if col: col_count[col - 1] += 1
    return col_count.index(max(col_count)) + 1

def columna_con_mas_paridad(hist, par=True):
    col_count = [0, 0, 0]
    for n in hist:
        if (int(n) % 2 == 0 and par) or (int(n) % 2 != 0 and not par):
            col = column_of_number(int(n))
            if col: col_count[col - 1] += 1
    return col_count.index(max(col_count)) + 1

def columna_por_transicion(numero, tabla):
    siguientes = tabla.get(numero, [])
    columnas = [column_of_number(n) for n in siguientes if column_of_number(n) is not None]
    return max(set(columnas), key=columnas.count) if columnas else None

def columna_reforzada(hist):
    ultimos = [column_of_number(int(n)) for n in hist[:6]]
    if ultimos.count(ultimos[0]) >= 4:
        return ultimos[0]
    return None
def columna_aprendida_de_errores(tabla, historial, window=20):
    errores = [fila for fila in tabla[-window:] if fila["Acierto"] == "NO"]
    if len(errores) < 5:
        return None  # No hay suficiente información para corregir

    columnas_que_salieron = []
    for fila in errores:
        try:
            n = int(fila["Numero que salio"])
            col = column_of_number(n)
            if col:
                columnas_que_salieron.append(col)
        except:
            continue

    if not columnas_que_salieron:
        return None

    sugerida = max(set(columnas_que_salieron), key=columnas_que_salieron.count)
    return sugerida
URL = "https://gamblingcounting.com/immersive-roulette"
# WINDOW_SIZE = 3
progresion = [1, 1, 2, 3, 4, 6, 9, 13, 19]
#progresion = [1, 1, 2, 3, 4, 6, 9, 13, 19, 28, 41, 60, 88, 129]
st.markdown("#### 📞 Versión 8.0 Optimizada")
valor_ficha = st.sidebar.number_input("💰 Valor de cada ficha (CLP)", min_value=100, max_value=100000, value=500, step=100)
saldo_inicial = st.sidebar.number_input("🏦 Saldo inicial del jugador (CLP)", min_value=1000, max_value=1000000, value=30000, step=1000)
if 'tiempo_inicio' not in st.session_state:
    st.session_state.tiempo_inicio = time.time()

# --- Funciones adicionales ---

# Coordenadas relativas (según imagen que compartiste)
COORDENADAS_COLUMNAS = {
    1: (0.74, 0.800), # --- OK COORDENADAS ---
    2: (0.74, 0.760), # --- OK COORDENADAS ---
    3: (0.74, 0.730), # --- OK COORDENADAS ---
}

def activar_pantalla_completa_selenium(driver):
    try:
        # Espera explícita de 5 segundos después de cargar la página
        time.sleep(15)

        # Espera a que el botón esté disponible y se pueda hacer clic
        WebDriverWait(driver, 15).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-role="fullscreen-button"]'))
        )
        boton = driver.find_element(By.CSS_SELECTOR, 'button[data-role="fullscreen-button"]')
        boton.click()
        print("🖥️ Botón de pantalla completa presionado con Selenium")
    except Exception as e:
        print("❌ No se pudo activar pantalla completa:", e)

def enviar_correo(resumen):
    remitente = "hyj.medina@gmail.com"  # ← tu correo Gmail
    destinatario = "hyj.medina@gmail.com"  # ← puede ser el mismo u otro
    asunto = "📈 Resumen Ruleta IA"
    contraseña_app = "hpsc nnnb uags ihwx"  # ← tu contraseña de aplicación de 16 caracteres

    mensaje = MIMEMultipart()
    mensaje["From"] = remitente
    mensaje["To"] = destinatario
    mensaje["Subject"] = asunto

    cuerpo = MIMEText(resumen, "plain")
    mensaje.attach(cuerpo)

    try:
        servidor = smtplib.SMTP("smtp.gmail.com", 587)
        servidor.starttls()
        servidor.login(remitente, contraseña_app)
        servidor.sendmail(remitente, destinatario, mensaje.as_string())
        servidor.quit()
        print("📧 Correo enviado con éxito.")
    except Exception as e:
        print("❌ Error al enviar correo:", e)

def detectar_posicion_ruleta():
    time.sleep(3)
    ventanas = gw.getWindowsWithTitle("1win")
    if not ventanas:
        ventanas = gw.getWindowsWithTitle("Immersive")
    if ventanas:
        ventana = ventanas[0]
        x, y = ventana.left, ventana.top
        w, h = ventana.width, ventana.height
        resultado = f"📏 Posición: ({x}, {y}) | Tamaño: {w}x{h}\n"
        for i, monitor in enumerate(get_monitors()):
            if monitor.x <= x < monitor.x + monitor.width and monitor.y <= y < monitor.y + monitor.height:
                resultado += f"🖥️ Monitor detectado: {i + 1} ({monitor.width}x{monitor.height})"
                break
        return resultado
    else:
        return "❌ No se detectó ventana activa de la ruleta."

def apostar_columna(columna, fichas=1):
    rel = COORDENADAS_COLUMNAS.get(columna)
    if not rel:
        return
    ventanas = gw.getWindowsWithTitle("1win")
    if not ventanas:
        ventanas = gw.getWindowsWithTitle("Immersive")
    if ventanas:
        ventana = ventanas[0]
        x_abs = ventana.left + int(ventana.width * rel[0])
        y_abs = ventana.top + int(ventana.height * rel[1])
        for _ in range(fichas):
            pyautogui.click(x_abs, y_abs)
            time.sleep(0.1)  # Pequeña pausa entre clics

# --- Modelo ML: entrenamiento y predicción ---
def preparar_datos_ml(historial, ventana=3):
    datos = []
    for i in range(len(historial) - ventana):
        fila = {}
        for j in range(ventana):
            n = int(historial[i + j])
            fila[f'n{j+1}'] = n
            fila[f'col{j+1}'] = column_of_number(n)
            fila[f'color{j+1}'] = obtener_color(n)
            fila[f'voisin{j+1}'] = 1 if n in VOISINS_DU_ZERO else 0
            fila[f'par{j+1}'] = 1 if n % 2 == 0 else 0
            fila[f'alto{j+1}'] = 1 if n >= 19 else 0
            fila[f'docena{j+1}'] = 1 if 1 <= n <= 12 else 2 if 13 <= n <= 24 else 3 if 25 <= n <= 36 else 0
        siguiente = int(historial[i + ventana])
        target = column_of_number(siguiente)
        if target is not None:
            fila['col_target'] = target
            datos.append(fila)
    return pd.DataFrame(datos)

def entrenar_modelo_rf(df, evaluar=False):
    X = df.drop('col_target', axis=1)
    y = df['col_target']
    cat_features = [col for col in X.columns if 'color' in col]
    num_features = [col for col in X.columns if col not in cat_features]

    preprocessor = ColumnTransformer([
        ('num', 'passthrough', num_features),
        ('cat', OneHotEncoder(handle_unknown='ignore'), cat_features)
    ])

    model = Pipeline([
        ('prep', preprocessor),
        ('clf', RandomForestClassifier(n_estimators=100, random_state=42, class_weight='balanced'))
    ])

    if evaluar:
        X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3, random_state=42)
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        acc = accuracy_score(y_test, y_pred)
        st.sidebar.markdown(f"📊 **Precisión del modelo (validación):** {acc*100:.1f}%")
        st.session_state.historial_accuracy.append(acc * 100)
        print("⚙️ Reporte de clasificación:\n", classification_report(y_test, y_pred))
    else:
        model.fit(X, y)

    return model

def predecir_con_confianza(modelo, fila):
    df_fila = pd.DataFrame([fila])  # ✅ Convertimos a DataFrame de una sola fila
    proba = modelo.predict_proba(df_fila)[0]
    predicho = modelo.classes_[np.argmax(proba)]
    confianza = max(proba)
    return predicho, confianza

# --- Heurística de respaldo (extendida) ---
def heuristica_columna():
    votos = []
    historial = st.session_state.historial[:50]
    transicion = construir_tabla_transicion(historial[:100])
    ultimo_num = int(historial[0])

    col1 = columna_con_mas_voisins(historial)
    col2 = columna_con_mas_altos(historial)
    col3 = columna_con_mas_bajos(historial)
    col4 = columna_por_transicion(ultimo_num, transicion)
    col5 = columna_con_mas_decenas(historial)
    col6 = columna_con_mas_paridad(historial, par=True)
    col7 = columna_con_mas_paridad(historial, par=False)
    col8 = columna_reforzada(historial)

    for col in [col1, col2, col3, col4, col5, col6, col7, col8]:
        if col:
            votos.append(col)

    # ➕ Agregar voto desde el modelo de aprendizaje de errores pasados
    col9 = columna_aprendida_de_errores(st.session_state.tabla_jugadas, historial)
    if col9:
        votos.append(col9)

    return max(set(votos), key=votos.count) if votos else 1

# --- Botones de control lateral ---
boton_activo = len(st.session_state.historial) > 0
st.sidebar.markdown("<br>", unsafe_allow_html=True)
st.sidebar.markdown("""
<div style='padding:1.5rem; margin-top:1.5rem; margin-bottom:1.5rem;
             background:#f5faff; border-left:6px solid #0099cc; border-radius:12px;
             box-shadow: 2px 2px 10px rgba(0,0,0,0.05);'>
<h4 style='margin-top:0; font-size:1.1rem;'>🎠 Control Ruleta</h4>
""", unsafe_allow_html=True)

ejecutar = st.sidebar.button("🚀 EJECUTAR RULETA", disabled=not boton_activo)
if ejecutar:
    try:
        subprocess.Popen(["python", "Robot_Ruleta_Apostador.py"])
        st.sidebar.success("✅ Ruleta abierta en Chrome.")
        st.session_state.resultado_ruleta = detectar_posicion_ruleta()
    except Exception as e:
        st.sidebar.error(f"❌ Error: {e}")

st.sidebar.markdown("</div>", unsafe_allow_html=True)

st.sidebar.markdown("<div style='margin-top:1.5rem;'></div>", unsafe_allow_html=True)

# 🎯 Botón para activar apuestas automáticas
st.sidebar.markdown("<div style='margin-top:1.5rem;'></div>", unsafe_allow_html=True)
st.sidebar.markdown("### ⚙️ Apuestas Automáticas")
activar_autoplay = st.sidebar.checkbox("🎯 Ejecutar Apuestas Automáticas")

# 📈 Historial de precisión del modelo
# if st.session_state.historial_accuracy:
#    st.sidebar.markdown("### 📈 Historial de precisión")
#    st.sidebar.line_chart(st.session_state.historial_accuracy)

st.sidebar.markdown("<div style='margin-top:1.5rem;'></div>", unsafe_allow_html=True)
if st.sidebar.button("🛑 Cerrar navegador"):
    try:
        st.session_state.driver.quit()
        st.success("✅ Navegador cerrado correctamente.")
    except:
        st.warning("⚠️ No se pudo cerrar el navegador.")

# --- Funciones restantes y lógica del modelo ---

def extraer_textos_validos(elems):
    return [e.text.strip() for e in elems if e.text.strip()]

def guardar_en_excel(numeros, archivo="Historial_Numeros_Ruleta.xlsx"):
    if not numeros:
        return

    ahora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ruta = Path(archivo)

    try:
        if not ruta.exists():
            wb = Workbook()
            ws = wb.active
            ws.title = "Historial"
            ws.append(["Numero", "Dia y hora de la jugada"])
        else:
            wb = load_workbook(archivo)
            ws = wb.active

        for numero in numeros:
            ws.insert_rows(idx=2)
            ws.cell(row=2, column=1, value=numero)
            ws.cell(row=2, column=2, value=ahora)

        wb.save(archivo)

        return True  # Se guardó con éxito

    except PermissionError:
        # Si el archivo está abierto, guardamos estos números en espera
        st.session_state.numeros_no_guardados.extend(numeros)
        return False

def generar_firma(lista):
    return hashlib.sha256("|".join(lista).encode("utf-8")).hexdigest()

def iniciar_driver():
    options = Options()
    options.add_argument("--headless")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--lang=es-CL")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("user-agent=Mozilla/5.0")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    driver.get(URL)
    # Espera y activa pantalla completa después de abrir la página
    activar_pantalla_completa_selenium(driver)
    return driver

def obtener_numeros_ruleta(driver):
    try:
        container_css = 'section[aria-labelledby="live-game-result-label"] .live-game-page__block__results--roulette.live-game-page__block__content'
        WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.CSS_SELECTOR, container_css)))
        container = driver.find_element(By.CSS_SELECTOR, container_css)
        small_selector = "div.roulette-number--small"
        WebDriverWait(container, 15).until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, small_selector)))
        for intento in range(3):
            try:
                elementos = container.find_elements(By.CSS_SELECTOR, small_selector)
                numeros = extraer_textos_validos(elementos)
                break
            except StaleElementReferenceException:
                time.sleep(1)
        else:
            raise Exception("No se pudo extraer elementos válidos tras varios intentos.")

        firma = generar_firma(numeros[:5])
        return numeros, firma

    except Exception as e:
        if "invalid session id" in str(e).lower() or "disconnected" in str(e).lower():
            st.warning("⚠️ Navegador desconectado. Reiniciando sesión...")
            try:
                driver.quit()
            except:
                pass
            nuevo_driver = iniciar_driver()
            st.session_state.driver = nuevo_driver
            return obtener_numeros_ruleta(nuevo_driver)
        else:
            st.error(f"Error al extraer datos: {e}")
            return [], None

def column_of_number(n):
    if n == 0: return None
    return 1 if n % 3 == 1 else 2 if n % 3 == 2 else 3

def columna_mas_frecuente(hist):
    frec = [0, 0, 0]
    for n in hist:
        col = column_of_number(int(n))
        if col: frec[col - 1] += 1
    return frec.index(max(frec)) + 1

def construir_tabla_transicion(historial, top_n=4):
    transiciones = {}
    for i in range(len(historial) - 1):
        actual = int(historial[i])
        siguiente = int(historial[i + 1])
        if actual not in transiciones:
            transiciones[actual] = []
        transiciones[actual].append(siguiente)

    # Convertimos a dict con top N más frecuentes
    tabla_final = {}
    for numero, siguientes in transiciones.items():
        conteo = pd.Series(siguientes).value_counts()
        tabla_final[numero] = conteo.head(top_n).index.tolist()
    
    return tabla_final

def top_n_mas_frecuentes(hist, n=10):
    conteo = pd.Series([int(x) for x in hist if x != '0']).value_counts()
    return set(conteo.head(n).index.tolist())

def obtener_color(n):
    try:
        ni = int(n)
        if ni == 0: return "green"
        if ni in NUMEROS_ROJOS: return "red"
        if ni in NUMEROS_NEGROS: return "black"
    except: pass
    return "gray"

def balancear_df(df):
    """Balancea el dataframe para que todas las clases de columna objetivo tengan igual cantidad."""
    min_count = df['col_target'].value_counts().min()
    df_balanceado = (
        df.groupby('col_target', group_keys=False)
        .apply(lambda x: x.sample(min_count, random_state=42))
    )
    return df_balanceado.reset_index(drop=True)

def mostrar_numeros(lista):
    colores = {
        'red': [1,3,5,7,9,12,14,16,18,19,21,23,25,27,30,32,34,36],
        'black': [2,4,6,8,10,11,13,15,17,20,22,24,26,28,29,31,33,35],
        'green': [0]
    }

    def get_color(n):
        if n in colores['red']:
            return 'red'
        elif n in colores['black']:
            return 'black'
        elif n in colores['green']:
            return 'green'
        else:
            return 'gray'

    bloques = ""
    for n in lista:
        try:
            val = int(n)
            color = get_color(val)
            bloques += f'<div class="num {color}">{val}</div>'
        except:
            pass

    html_code = f"""
    <style>
    .grid {{
        display: flex;
        flex-wrap: wrap;
        max-height: 400px;
        overflow-y: auto;
        padding: 10px;
        border: 1px solid #ccc;
        border-radius: 8px;
        background: #f9f9f9;
    }}
    .num {{
        width: 30px;
        height: 30px;
        margin: 3px;
        display: flex;
        justify-content: center;
        align-items: center;
        font-weight: bold;
        font-size: 16px;
        border-radius: 4px;
        color: white;
    }}
    .red {{ background-color: red; }}
    .black {{ background-color: black; }}
    .green {{ background-color: green; }}
    .gray {{ background-color: gray; }}
    </style>

    <div class="grid">
        {bloques}
    </div>
    """
    components.html(html_code, height=450)


def obtener_indice_progresion():
    # ➤ Si estamos en reset, devolvemos 0 y apagamos el flag
    if st.session_state.reset_progression:
        st.session_state.reset_progression = False
        return 0

    # Código original: cuenta fallos hasta el último acierto
    indice = 0
    for jugada in reversed(st.session_state.tabla_jugadas):
        if jugada["Acierto"] == "NO":
            indice += 1
        else:
            break
    return indice

  
# --- TAB 1: Modelo Predictivo ---
with tab1:
    if not st.session_state.initialized:
        st.session_state.tiempo_inicio = datetime.now()
        st.session_state.driver = iniciar_driver()
        nums, firma = obtener_numeros_ruleta(driver=st.session_state.driver)
        if nums and firma:
            for n in reversed(nums):
                st.session_state.historial.insert(0, n)
                st.session_state.fechas.insert(0, datetime.now())
                st.session_state.data_arr = np.append(st.session_state.data_arr, int(n))
            st.session_state.total_inicial = len(nums)
            st.session_state.ultima_firma = firma
            st.session_state.last_col = columna_mas_frecuente(st.session_state.historial[:70]) if len(st.session_state.historial) >= 70  else column_of_number(int(nums[0]))

            guardar_en_excel(reversed(nums))  # 👈 Aquí se ejecuta ANTES de reiniciar la app
        st.session_state.initialized = True
        st.rerun()

    col1, col2 = st.columns([2, 1])

    with col2:
        if not st.session_state.pausar and st.session_state.tabla_jugadas:
            indice_prog_tmp = obtener_indice_progresion()
            fichas_apostadas = progresion[min(indice_prog_tmp, len(progresion)-1)]
            ultimo_numero = st.session_state.historial[0] if st.session_state.historial else "?"

            # --- Mostrar resumen de la próxima jugada ---
            st.markdown(f"""
            <div style='padding:1rem;background:#f0fcfc;border-radius:10px;border-left:5px solid #70c4c4;'>
            🟢 <strong>Próxima apuesta:</strong><br>
            🎯 <strong>Columna {st.session_state.last_col}</strong><br>
            🎲 <strong>{fichas_apostadas} fichas</strong> (= ${fichas_apostadas * valor_ficha:,} CLP)<br>
            🔢 <strong>N° de jugada:</strong> {len(st.session_state.tabla_jugadas) + 1}<br>
            📍 <strong>Apostar después del número:</strong> {ultimo_numero}<br>
            </div>
            """, unsafe_allow_html=True)

            st.markdown("<div style='margin-bottom: 1.5rem;'></div>", unsafe_allow_html=True)

            df = pd.DataFrame(st.session_state.tabla_jugadas)
            df = df.drop(columns=["Perdida", "% de efectividad"], errors='ignore')

            mostrar_todas = st.checkbox("🔎 Mostrar todas las jugadas", value=False)
            if mostrar_todas:
                st.dataframe(df, use_container_width=True)
            else:
                st.dataframe(df.tail(20), use_container_width=True)

            # 🧠 Mostrar desglose de votos heurísticos (solo si hay historial suficiente)
            if len(st.session_state.historial) >= 10:
                historial = st.session_state.historial[:50]
                transicion = construir_tabla_transicion(historial[:100])
                ultimo_num = int(historial[0])

                votos_heuristica = {}

                def contar_voto(nombre, columna):
                    if columna:
                        votos_heuristica[nombre] = f"Columna {columna}"

                contar_voto("Voisins", columna_con_mas_voisins(historial))
                contar_voto("Altos", columna_con_mas_altos(historial))
                contar_voto("Bajos", columna_con_mas_bajos(historial))
                contar_voto("Transición", columna_por_transicion(ultimo_num, transicion))
                contar_voto("Decenas", columna_con_mas_decenas(historial))
                contar_voto("Pares", columna_con_mas_paridad(historial, par=True))
                contar_voto("Impares", columna_con_mas_paridad(historial, par=False))
                contar_voto("Columna Reforzada", columna_reforzada(historial))
                contar_voto("Corrección por errores", columna_aprendida_de_errores(st.session_state.tabla_jugadas, historial))

                st.markdown("### 🧩 Votación heurística")
                for heuristica, col in votos_heuristica.items():
                    st.markdown(f"- **{heuristica}:** {col}")

    with col1:
        total = st.session_state.total_inicial + st.session_state.spins_nuevos
        total_jugadas = len(st.session_state.tabla_jugadas)

        # Calcular % de ciclos exitosos (progresión Martingala de hasta 8  pasos)
        aciertos_lista = [fila["Acierto"] for fila in st.session_state.tabla_jugadas]
        max_pasos = 8
        total_ciclos = 0
        ciclos_exitosos = 0
        i = 0

        while i < len(aciertos_lista):
            ciclo = aciertos_lista[i:i + max_pasos]
            total_ciclos += 1
            if "SI" in ciclo:
                ciclos_exitosos += 1
                i += ciclo.index("SI") + 1
            else:
                i += max_pasos

        porcentaje_ciclos = f"{(ciclos_exitosos / total_ciclos * 100):.1f}%" if total_ciclos > 0 else "0%"

        # Calcular ganancia acumulada
        saldo_tmp = saldo_inicial
        indice_prog_tmp = 0

        for jugada in st.session_state.tabla_jugadas:
            if jugada["Columna Predecida"] == "Sin apostar":
                continue  # 👈 saltamos la fila inicial sin apuesta

            if indice_prog_tmp >= len(progresion):
                break

            fichas_apostadas = progresion[indice_prog_tmp]
            monto_apostado = fichas_apostadas * valor_ficha

            if jugada["Acierto"] == "SI":
                saldo_tmp += monto_apostado * 2
                indice_prog_tmp = 0
            else:
                saldo_tmp -= monto_apostado
                indice_prog_tmp += 1

        # ganancia_real = saldo_tmp - saldo_inicial
        st.markdown(f"🌀 **Total de Jugadas:** {total_jugadas}")
        st.markdown(f"📊 **Efectividad:** {porcentaje_ciclos}")

        # Cálculo de ganancia + tiempo en formato resumido
        ganancia_real = saldo_tmp - saldo_inicial
        tiempo_total = time.time() - st.session_state.tiempo_inicio.timestamp()
        horas = int(tiempo_total // 3600)
        minutos = int((tiempo_total % 3600) // 60)
        segundos = int(tiempo_total % 60)

        st.markdown(f"""
        <div style='padding:1rem; background:#f0f0f0; border-left:5px solid #009933;
                    border-radius:10px; margin-top:0.5rem; font-size:16px;'>
        💰 <strong>Ganancia Real del Sistema:</strong> ${ganancia_real:,} CLP en {horas:02d}:{minutos:02d}:{segundos:02d} hrs
        </div>
        """, unsafe_allow_html=True)

        # --- Enviar resumen por correo cada 5 jugadas una sola vez ---
        total_jugadas = len(st.session_state.tabla_jugadas)

        if total_jugadas % 5 == 0 and total_jugadas > 0:
            if st.session_state.correo_enviado_ultima_vez != total_jugadas:
                resumen = (
                    f"🎰 Estrategia Segura - Ruleta IA\n\n"
                    f"Jugadas: {total_jugadas}\n"
                    f"Ganancia acumulada: ${ganancia_real:,} CLP\n"
                    f"Fecha y hora: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                )
                enviar_correo(resumen)
                st.session_state.correo_enviado_ultima_vez = total_jugadas

        # Mostrar los siguientes números más probables según transición (con color)
        if st.session_state.historial:
            ultimo = int(st.session_state.historial[0])
            siguientes = st.session_state.matriz_transicion.get(ultimo, [])
            if siguientes:
                bloques_html = ""
                for n in siguientes:
                    color = obtener_color(n)
                    bloques_html += f'<div class="num {color}">{n}</div>'

                html_code = f"""
                <style>
                .transicion-grid {{
                    display: flex;
                    gap: 5px;
                    padding: 10px 0;
                }}
                .num {{
                    width: 30px;
                    height: 30px;
                    display: flex;
                    justify-content: center;
                    align-items: center;
                    font-weight: bold;
                    font-size: 16px;
                    border-radius: 4px;
                    color: white;
                }}
                .red {{ background-color: red; }}
                .black {{ background-color: black; }}
                .green {{ background-color: green; }}
                .gray {{ background-color: gray; }}
                </style>

                <div><strong>🎯 Probables siguientes números tras el {ultimo}:</strong></div>
                <div class="transicion-grid">
                    {bloques_html}
                </div>
                """
                components.html(html_code, height=70)

        mostrar_numeros(st.session_state.historial[:total])
        st.markdown("<div style='margin-bottom:1.5rem'></div>", unsafe_allow_html=True)

        espera = st.empty()
        for i in range(6, 0, -1): # <-- aquí defines cuántos segundos esperar
            espera.info(f"🌀 Esperando próximo número en {i}s...")
            time.sleep(1)

        # --- OBTENER NUEVO NÚMERO ---
        nums2, firma2 = obtener_numeros_ruleta(driver=st.session_state.driver)
        if nums2 and firma2 and firma2 != st.session_state.ultima_firma:
            nuevo = nums2[0]
            st.session_state.historial.insert(0, nuevo)
            st.session_state.fechas.insert(0, datetime.now())
            st.session_state.ultima_firma = firma2
            st.session_state.spins_nuevos += 1
            guardar_en_excel([nuevo])  # Guarda el nuevo número con su hora

            # --- EVALUAR ACUERDO A PREDICCIÓN ANTERIOR ---
            if st.session_state.columna_predicha_anterior is not None:
                col_actual = column_of_number(int(nuevo))
                acierto = (col_actual == st.session_state.columna_predicha_anterior)
                st.session_state.aciertos += int(acierto)
                st.session_state.perdidas += int(not acierto)
                jugada = len(st.session_state.tabla_jugadas) + 1

                fichas_usadas = progresion[min(obtener_indice_progresion(), len(progresion)-1)]
                dinero_apostado = fichas_usadas * valor_ficha

                st.session_state.tabla_jugadas.append({
                    "Jugada": jugada,
                    "Columna Predecida": f"Columna {st.session_state.columna_predicha_anterior}",
                    "Fichas Apostadas": fichas_usadas,
                    "Dinero Apostado": dinero_apostado,
                    "Numero que salio": nuevo,
                    "Acierto": "SI" if acierto else "NO"
                })

                # ← Aquí insertas este bloque:
                if fichas_usadas >= 3000 and not acierto and st.session_state.simular_restantes == 0:
                    st.session_state.simular_restantes = 2
                    st.session_state.pausar = True
                    st.sidebar.warning(
                        "🚨 Límite alcanzado: detengo apuesta automática.\n"
                        "   Simularé las próximas 2 jugadas y luego resetearé progresión."
                    )

                # 👇 Ejecuta script para mantener la pantalla activa
                subprocess.Popen(["python", "Mantener_Activo_PC.py"])

            # --- NUEVA PREDICCIÓN (votación combinada con ciclo de Voisins) ---
            transicion = st.session_state.matriz_transicion
            ultimo_num = int(nuevo)

            def columna_por_transicion(numero, tabla):
                siguientes = tabla.get(numero, [])
                if not siguientes:
                    return None
                columnas = [column_of_number(n) for n in siguientes if column_of_number(n) is not None]
                if not columnas:
                    return None
                return max(set(columnas), key=columnas.count)

            NUMEROS_ALTOS = set(range(19, 37))
            NUMEROS_BAJOS = set(range(1, 19))

            def columna_con_mas_voisins(historial):
                col_count = [0, 0, 0]
                for n in historial:
                    if int(n) in VOISINS_DU_ZERO:
                        col = column_of_number(int(n))
                        if col: col_count[col - 1] += 1
                return col_count.index(max(col_count)) + 1

            def columna_con_mas_altos(historial):
                col_count = [0, 0, 0]
                for n in historial:
                    if int(n) in NUMEROS_ALTOS:
                        col = column_of_number(int(n))
                        if col: col_count[col - 1] += 1
                return col_count.index(max(col_count)) + 1

            def columna_con_mas_bajos(historial):
                col_count = [0, 0, 0]
                for n in historial:
                    if int(n) in NUMEROS_BAJOS:
                        col = column_of_number(int(n))
                        if col: col_count[col - 1] += 1
                return col_count.index(max(col_count)) + 1

            def se_espera_voisin(historial, max_intervalo=6):
                posiciones = [i for i, n in enumerate(historial) if int(n) in VOISINS_DU_ZERO]
                if len(posiciones) < 3:
                    return False
                intervalos = [posiciones[i] - posiciones[i+1] for i in range(len(posiciones)-1)]
                promedio = sum(intervalos) / len(intervalos)
                desde_ultimo = posiciones[0]
                return desde_ultimo >= promedio - 1 and desde_ultimo <= promedio + 1

            # --- Cálculo de las 4 columnas base (agregado números altos)
            col1 = columna_mas_frecuente(st.session_state.historial[:100])
            col2 = columna_con_mas_voisins(st.session_state.historial[:100])
            col3 = columna_por_transicion(ultimo_num, transicion)
            col4 = columna_con_mas_altos(st.session_state.historial[:100])
            col5 = columna_con_mas_bajos(st.session_state.historial[:100])

            # --- Votación base (agregando columna por altos)
            votos = [col for col in [col1, col2, col3, col4, col5] if col is not None]

            # --- Resultado final
            # --- Resultado final (ML + Heurística híbrida)
            if len(st.session_state.historial) >= 180:
                st.session_state.contador_ciclos += 1

                # Solo reentrenamos cada 10 ciclos
                if st.session_state.contador_ciclos % 10 == 0 or st.session_state.modelo_rf_largo is None:
                    df_entrenamiento = preparar_datos_ml(st.session_state.historial[:130])
                    df_entrenamiento = balancear_df(df_entrenamiento)  # 👈 Balanceamos clases antes de entrenar
                    st.session_state.modelo_rf_largo = entrenar_modelo_rf(df_entrenamiento, evaluar=True)

                modelo_rf = st.session_state.modelo_rf_largo

                # Prepara entrada de predicción
                entrada = {}
                for j in range(3):
                    n = int(st.session_state.historial[j])
                    entrada[f'n{j+1}'] = n
                    entrada[f'col{j+1}'] = column_of_number(n)
                    entrada[f'color{j+1}'] = obtener_color(n)
                    entrada[f'voisin{j+1}'] = 1 if n in VOISINS_DU_ZERO else 0
                    entrada[f'par{j+1}'] = 1 if n % 2 == 0 else 0
                    entrada[f'alto{j+1}'] = 1 if n >= 19 else 0
                    entrada[f'docena{j+1}'] = 1 if 1 <= n <= 12 else 2 if 13 <= n <= 24 else 3 if 25 <= n <= 36 else 0

                pred_col_ml, conf_ml = predecir_con_confianza(modelo_rf, entrada)

# --------------------------------------------------------------------------------------

                col_heuristica = max(set(votos), key=votos.count)

# --------------------------------------------------------------------------------------

                if conf_ml >= 0.55:
                    nueva_columna = pred_col_ml
                else:
                    nueva_columna = col_heuristica


            # --- GUARDAR columna predicha para la siguiente jugada ---
            st.session_state.last_col = nueva_columna
            st.session_state.columna_predicha_anterior = nueva_columna

            # Decide si apostamos de verdad o estamos en simulación
            # ————— 1) Cuenta atrás de simulación (siempre) —————
            if st.session_state.simular_restantes > 0:
                st.session_state.simular_restantes -= 1
                st.sidebar.info(f"🔍 Simulando jugadas restantes: {st.session_state.simular_restantes}")
                if st.session_state.simular_restantes == 0:
                    st.session_state.reset_progression = True
                    st.sidebar.success("♻️ Simulación terminada: progresión reiniciada a 500.")

            # ————— 2) Apuesta real (solo si ya terminó simulación y está activado) —————
            elif activar_autoplay:
                indice_prog_tmp = obtener_indice_progresion()
                fichas_usadas   = progresion[min(indice_prog_tmp, len(progresion)-1)]
                apostar_columna(nueva_columna, fichas=fichas_usadas)

            # --- ACTUALIZAR MATRIZ DE TRANSICIÓN ---
            if len(st.session_state.historial) > 25:
                st.session_state.matriz_transicion = construir_tabla_transicion(st.session_state.historial[:50])

            # --- MANEJAR CAMBIO DE COLUMNA POR FALLOS ---
            errores = obtener_indice_progresion()
            if errores >= 6:
                st.error("""
                🚨 **RULETA MONITOREADA DESDE EL CASINO**

                Se han detectado **jugadas consecutivas sin acierto**.  
                Esto podría indicar un patrón manipulado o vigilancia del sistema.

                🕒 **Recomendación:** Detener la estrategia al menos por **30 minutos** y reanudar con análisis limpio.
                """)
                st.session_state.pausar = True
            elif errores >= 5:
                col_freq = columna_mas_frecuente(st.session_state.historial[:50])
                if st.session_state.last_col == col_freq:
                    st.session_state.pausar = True
                else:
                    st.session_state.last_col = col_freq

            if st.session_state.pausar:
                st.warning("🔁 Estrategia pausada por baja efectividad. Esperando nuevo patrón...")
                st.session_state.pausar = False

# --- Recarga automática ---
time.sleep(1)
if st.session_state.numeros_no_guardados:
    guardado = guardar_en_excel(st.session_state.numeros_no_guardados)
    if guardado:
        st.session_state.numeros_no_guardados = []
st.rerun()